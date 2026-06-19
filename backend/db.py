$path = "backend\db.py"
$content = Get-Content $path -Raw

$marker = "    # Create/repair default admin using runtime bcrypt so hash is always correct."
$before = $content.Substring(0, $content.IndexOf($marker))

$replacement = @'
    # Seed settings table
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS app_settings (
            setting_key TEXT PRIMARY KEY,
            setting_value TEXT NOT NULL
        )
        """
    )

    def get_setting(key):
        cur.execute("SELECT setting_value FROM app_settings WHERE setting_key=?", (key,))
        row = cur.fetchone()
        return row["setting_value"] if row else None

    def set_setting(key, value):
        cur.execute(
            """
            INSERT INTO app_settings (setting_key, setting_value)
            VALUES (?, ?)
            ON CONFLICT(setting_key) DO UPDATE SET setting_value=excluded.setting_value
            """,
            (key, value),
        )

    # Seed admin login only once
    try:
        from services.password_service import hash_password

        if get_setting("admin_seed_done") != "1":
            admin_hash = hash_password("Admin@123")

            cur.execute("SELECT id FROM clients WHERE client_code=?", ("admin",))
            if cur.fetchone():
                cur.execute(
                    """
                    UPDATE clients
                    SET name=?, mobile=?, password_hash=?, role=?, status=?, must_change_password=0
                    WHERE client_code=?
                    """,
                    ("Admin User", "0000000000", admin_hash, "admin", "active", "admin"),
                )
            else:
                cur.execute(
                    """
                    INSERT INTO clients
                    (client_code, name, mobile, password_hash, role, status, must_change_password)
                    VALUES (?, ?, ?, ?, ?, ?, 0)
                    """,
                    ("admin", "Admin User", "0000000000", admin_hash, "admin", "active"),
                )

            set_setting("admin_seed_done", "1")
            print("[database] Admin login ready: admin / Admin@123")

    except Exception as exc:
        print(f"[database] Could not seed admin: {exc}")

    # Seed initial 100 clients permanently from backend/seed/client_master_100_sample.xlsx
    try:
        from openpyxl import load_workbook
        from services.password_service import hash_password
        from datetime import datetime, date

        def clean_cell(value):
            if value is None:
                return ""
            if isinstance(value, (datetime, date)):
                return value.strftime("%Y-%m-%d")
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
            return str(value).strip()

        if get_setting("initial_100_clients_seed_done") != "1":
            seed_path = os.path.join(
                os.path.dirname(__file__),
                "seed",
                "client_master_100_sample.xlsx",
            )

            if os.path.exists(seed_path):
                wb = load_workbook(seed_path)
                ws = wb.active

                headers = [clean_cell(cell.value) for cell in ws[1]]
                header_map = {name: idx for idx, name in enumerate(headers)}

                required_cols = ["ClientCode", "Name", "Mobile"]
                missing = [col for col in required_cols if col not in header_map]

                if missing:
                    print(f"[database] Seed file missing columns: {missing}")
                else:
                    inserted = 0
                    updated = 0

                    for row in ws.iter_rows(min_row=2, values_only=True):
                        client_code = clean_cell(row[header_map["ClientCode"]])
                        name = clean_cell(row[header_map["Name"]])
                        mobile = clean_cell(row[header_map["Mobile"]])

                        if not client_code or not name or not mobile:
                            continue

                        pan = clean_cell(row[header_map["PAN"]]) if "PAN" in header_map else ""
                        dob = clean_cell(row[header_map["DOB"]]) if "DOB" in header_map else ""
                        parent_code = clean_cell(row[header_map["ParentCode"]]) if "ParentCode" in header_map else "admin"
                        status = clean_cell(row[header_map["Status"]]) if "Status" in header_map else "active"

                        if not parent_code:
                            parent_code = "admin"

                        status = status.lower()
                        if status not in ["active", "inactive"]:
                            status = "active"

                        cur.execute("SELECT id FROM clients WHERE client_code=?", (client_code,))
                        existing = cur.fetchone()

                        if existing:
                            cur.execute(
                                """
                                UPDATE clients
                                SET name=?,
                                    pan=?,
                                    dob=?,
                                    mobile=?,
                                    parent_code=?,
                                    role='client',
                                    status=?,
                                    must_change_password=0,
                                    updated_at=CURRENT_TIMESTAMP
                                WHERE client_code=?
                                """,
                                (name, pan, dob, mobile, parent_code, status, client_code),
                            )
                            updated += 1
                        else:
                            password_hash = hash_password(mobile)
                            cur.execute(
                                """
                                INSERT INTO clients
                                (client_code, name, pan, dob, mobile, parent_code,
                                 password_hash, role, status, must_change_password)
                                VALUES (?, ?, ?, ?, ?, ?, ?, 'client', ?, 0)
                                """,
                                (client_code, name, pan, dob, mobile, parent_code, password_hash, status),
                            )
                            inserted += 1

                    set_setting("initial_100_clients_seed_done", "1")
                    print(f"[database] Initial 100 clients seed done. Inserted={inserted}, Updated={updated}")

            else:
                print(f"[database] Seed Excel not found: {seed_path}")

    except Exception as exc:
        print(f"[database] Could not seed initial 100 clients: {exc}")

    conn.commit()
'@

Set-Content $path ($before + $replacement)